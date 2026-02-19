from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.utils import timezone
from django.db.models import Q
import openpyxl
from decimal import Decimal
from django.core.paginator import Paginator
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT


from .models import Product, Offer, OfferItem, ActivityLog, Notification
from .permissions_helpers import (
    can_user_send_offer,
    can_user_approve_offer,
    can_user_apply_discount,
    can_user_view_offer,
    can_user_edit_offer,
    can_user_delete_offer,
    check_offer_amount_threshold,
    log_activity,
    create_notification,
    notify_manager_for_approval,
    notify_user_on_manager_approval,
    notify_on_offer_status_change,
)
from decimal import Decimal
from django.contrib import messages

from django.core.exceptions import PermissionDenied
from functools import wraps
from django.http import Http404
from django.db import models
from accounts.models import User
from django.http import JsonResponse
from django.views.decorators.http import require_POST

# ============================================
# YETKİ KONTROL HELPER FONKSİYONLARI
# ============================================

def can_user_view_offer(user, offer):
    """Kullanıcı bu teklifi görüntüleyebilir mi?"""
    
    # Admin/Superadmin her şeyi görebilir
    if user.is_superuser or user.is_staff:
        return True
    
    # Kendi teklifi
    if offer.user == user:
        return True
    
    # Yönetici kendi firmasının/eczanesinin tekliflerini görebilir
    if user.is_manager and offer.user.role == user.role:
        # Aynı firma/eczane mi kontrol et
        if user.role == 'firma' and offer.user.company_name == user.company_name:
            return True
        elif user.role == 'eczane' and offer.user.pharmacy_name == user.pharmacy_name:
            return True
    
    # "Tüm teklifleri görebilir" yetkisi
    if hasattr(user, 'can_view_all_offers') and user.can_view_all_offers:
        if offer.user.role == user.role:
            return True
    
    return False, "Bu teklifi görüntüleme yetkiniz yok."


def can_user_edit_offer(user, offer):
    """Kullanıcı bu teklifi düzenleyebilir mi?"""
    
    # Admin/Superadmin her şeyi düzenleyebilir
    if user.is_superuser or user.is_staff:
        return True
    
    # Gönderilmiş/onaylanmış/reddedilmiş teklifler düzenlenemez
    if offer.status in ['sent', 'approved', 'rejected']:
        return False
    
    # Kendi teklifi + "kendi tekliflerini düzenleyebilir" yetkisi
    if offer.user == user:
        if hasattr(user, 'can_edit_own_offer') and user.can_edit_own_offer:
            return True
    
    # Yönetici veya "tüm teklifleri düzenleyebilir" yetkisi
    if user.is_manager or (hasattr(user, 'can_edit_all_offers') and user.can_edit_all_offers):
        if offer.user.role == user.role:
            return True
    
    return False


def can_user_delete_offer(user, offer):
    """Kullanıcı bu teklifi silebilir mi?"""
    
    # Admin/Superadmin her şeyi silebilir
    if user.is_superuser or user.is_staff:
        return True
    
    # Onaylanmış teklifler silinemez
    if offer.status == 'approved':
        return False
    
    # Kendi teklifi + "kendi tekliflerini silebilir" yetkisi
    if offer.user == user:
        if hasattr(user, 'can_delete_own_offer') and user.can_delete_own_offer:
            return True
    
    # Yönetici veya "tüm teklifleri silebilir" yetkisi
    if user.is_manager or (hasattr(user, 'can_delete_all_offers') and user.can_delete_all_offers):
        if offer.user.role == user.role:
            return True
    
    return False


def can_user_send_offer(user, offer):
    """Kullanıcı bu teklifi gönderebilir mi?"""
    
    # Admin/Superadmin her şeyi gönderebilir
    if user.is_superuser or user.is_staff:
        return True
    
    # Sadece draft teklifler gönderilebilir
    if offer.status != 'draft':
        return False
    
    # "Teklif gönderebilir" yetkisi olmalı
    if not (hasattr(user, 'can_send_offer') and user.can_send_offer):
        return False
    
    # Kendi teklifi veya yönetici
    if offer.user == user or user.is_manager:
        return True
    
    return False


def can_user_approve_high_value_offer(user, offer):
    """Kullanıcı yüksek tutarlı teklifi onaylayabilir mi?"""
    
    # Admin/Superadmin her şeyi onaylayabilir
    if user.is_superuser or user.is_staff:
        return True
    
    # Yönetici + "50K+ onaylayabilir" yetkisi
    if user.is_manager and hasattr(user, 'can_approve_high_value_offers'):
        if user.can_approve_high_value_offers:
            return True
    
    return False

def role_required(role_name):
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped(request, *args, **kwargs):
            if not request.user.is_authenticated:
                raise PermissionDenied
            
            if request.user.is_superuser:
                return view_func(request, *args, **kwargs)
            
            if getattr(request.user, "role", None) != role_name:
                raise PermissionDenied
            
            return view_func(request, *args, **kwargs)
        return _wrapped
    return decorator


# =======================
# ÜRÜN LİSTESİ
# =======================
from django.core.paginator import Paginator

def product_list(request):
    all_products = Product.objects.all().order_by('name')
    
    search_query = request.GET.get('search', '').strip()
    if search_query:
        all_products = all_products.filter(
            models.Q(name__icontains=search_query) | 
            models.Q(barcode__icontains=search_query)
        )
    
    paginator = Paginator(all_products, 50)
    page_number = request.GET.get('page', 1)
    products = paginator.get_page(page_number)
    
    # Sepet bilgileri
    cart_count = 0
    cart_items = []
    cart_total = 0
    
    if request.user.is_authenticated:
        draft_offer = Offer.objects.filter(user=request.user, status='draft').first()
        if draft_offer:
            cart_count = draft_offer.items.count()
            cart_items = draft_offer.items.all()[:5]
            cart_total = draft_offer.gross_total_price()
    
    return render(request, "products/product_list.html", {
        "products": products,
        "search_query": search_query,
        "total_count": all_products.count(),
        "cart_count": cart_count,
        "cart_items": cart_items,
        "cart_total": cart_total,
    })

# =======================
# TEKLİFİM
# =======================
@login_required
def offer_view(request):
    offer = Offer.objects.filter(user=request.user, status="draft").order_by("-created_at").first()
    if offer is None:
        offer = Offer.objects.create(user=request.user, status="draft")
    return render(request, "products/offer.html", {"offer": offer})


# =======================
# TEKLİFE ÜRÜN EKLE
# =======================
@login_required
def add_to_offer(request, product_id):
    if request.method != "POST":
        return redirect("product_list")

    product = get_object_or_404(Product, id=product_id)

    qty_raw = request.POST.get("quantity", "1")
    try:
        quantity = int(qty_raw)
    except (TypeError, ValueError):
        quantity = 1
    if quantity < 1:
        quantity = 1

    offer = Offer.objects.filter(user=request.user, status="draft").order_by("-created_at").first()
    
    if offer is None:
        offer = Offer.objects.create(user=request.user, status="draft")

    item, created = OfferItem.objects.get_or_create(
        offer=offer,
        product=product,
        defaults={
            "quantity": quantity,
            "unit_price": product.price,
            "vat_rate": product.vat_rate,
        },
    )

    if not created:
        item.quantity += quantity
        item.save()

    return redirect(request.META.get('HTTP_REFERER', 'product_list'))


# =======================
# TEKLİFİ ECZANEYE GÖNDER
# =======================
@login_required
def send_offer_to_pharmacy(request):
    """Teklifi eczaneye gönder veya yönetici onayına sun"""
    offer = Offer.objects.filter(user=request.user, status="draft").order_by("-created_at").first()
    if offer is None:
        messages.error(request, "Gönderilecek teklif bulunamadı.")
        return redirect("offer")

    if request.method == "POST":
        # Yetki kontrolü
        if not can_user_send_offer(request.user, offer):
            # 50K üzeri ise yönetici onayına gönder
            if check_offer_amount_threshold(offer):
                offer.requires_manager_approval = True
                offer.manager_approval_pending = True
                offer.status = "draft"  # Draft olarak kalır
                offer.save()
                
                # Log kaydet
                log_activity(
                    user=request.user,
                    action='offer_sent_for_approval',
                    description=f"#{offer.id} nolu teklif ({offer.gross_total_price():,.2f} TL) yönetici onayına gönderildi.",
                    offer=offer,
                    request=request
                )
                
                # Yöneticiye bildirim gönder
                notify_manager_for_approval(offer)
                
                messages.warning(request, f"Teklif tutarı 50.000 TL'nin üzerinde olduğu için yönetici onayına gönderildi.")
                return redirect("my_offers")
            else:
                messages.error(request, "Teklif gönderme yetkiniz yok.")
                return redirect("offer")
        
        # Yetki varsa direkt gönder
        offer.status = "sent"
        offer.sent_at = timezone.now()
        offer.save()
        
        # Log kaydet
        log_activity(
            user=request.user,
            action='offer_sent',
            description=f"#{offer.id} nolu teklif eczaneye gönderildi. Tutar: {offer.gross_total_price():,.2f} TL",
            offer=offer,
            request=request
        )
        
        messages.success(request, "Teklif eczaneye gönderildi!")
    
    return redirect("my_offers")


# =======================
# YÖNETİCİ ONAY SİSTEMİ
# =======================
@role_required("firma")
def manager_approve_offer(request, offer_id):
    """Yönetici teklifi onaylar ve eczaneye gönderir"""
    offer = get_object_or_404(Offer, id=offer_id)
    
    # Sadece yönetici onaylayabilir
    if not request.user.is_manager:
        messages.error(request, "Bu işlem için yönetici yetkisi gereklidir.")
        return redirect("my_offers")
    
    # Onay bekleyen teklif mi?
    if not offer.manager_approval_pending:
        messages.error(request, "Bu teklif onay beklemiyro.")
        return redirect("my_offers")
    
    if request.method == "POST":
        # Onay
        offer.manager_approval_pending = False
        offer.approved_by_manager = request.user
        offer.manager_approved_at = timezone.now()
        offer.status = "sent"
        offer.sent_at = timezone.now()
        offer.save()
        
        # Log kaydet
        log_activity(
            user=request.user,
            action='manager_approved_offer',
            description=f"#{offer.id} nolu teklif yönetici tarafından onaylandı ve eczaneye gönderildi.",
            offer=offer,
            target_user=offer.user,
            request=request
        )
        
        # Personele bildirim
        notify_user_on_manager_approval(offer, approved=True)
        
        messages.success(request, f"Teklif onaylandı ve eczaneye gönderildi!")
        return redirect("my_offers")
    
    return redirect("my_offers")


@role_required("firma")
def manager_reject_offer(request, offer_id):
    """Yönetici teklifi reddeder"""
    offer = get_object_or_404(Offer, id=offer_id)
    
    # Sadece yönetici reddedebilir
    if not request.user.is_manager:
        messages.error(request, "Bu işlem için yönetici yetkisi gereklidir.")
        return redirect("my_offers")
    
    # Onay bekleyen teklif mi?
    if not offer.manager_approval_pending:
        messages.error(request, "Bu teklif onay beklemiyor.")
        return redirect("my_offers")
    
    if request.method == "POST":
        reason = request.POST.get('reason', '').strip()
        
        # Red
        offer.manager_approval_pending = False
        offer.manager_rejection_reason = reason
        offer.status = "draft"  # Draft'a geri döner
        offer.requires_manager_approval = False
        offer.save()
        
        # Log kaydet
        log_activity(
            user=request.user,
            action='manager_rejected_offer',
            description=f"#{offer.id} nolu teklif yönetici tarafından reddedildi. Sebep: {reason}",
            offer=offer,
            target_user=offer.user,
            request=request
        )
        
        # Personele bildirim
        notify_user_on_manager_approval(offer, approved=False)
        
        messages.warning(request, "Teklif reddedildi.")
        return redirect("my_offers")
    
    return redirect("my_offers")


# =======================
# ECZANE – GELEN TEKLİFLER
# =======================
@role_required("eczane")
def pharmacy_inbox(request):
    sent_offers = Offer.objects.filter(status="sent")
    
    revised_original_ids = sent_offers.filter(
        original_offer__isnull=False
    ).values_list('original_offer_id', flat=True).distinct()
    
    offers = sent_offers.exclude(
        id__in=revised_original_ids
    ).order_by("-sent_at")

    return render(
        request,
        "products/pharmacy_inbox.html",
        {"offers": offers}
    )


# =======================
# ECZANE – TEKLİF DETAY
# =======================
@role_required("eczane")
def pharmacy_offer_detail(request, offer_id):
    """Teklif detayı - Eczane tarafı"""
    offer = get_object_or_404(Offer, id=offer_id)
    
    # Firma adreslerini getir (onaylanan teklifler için teslimat adresi seçimi)
    firm_addresses = None
    if offer.status == 'approved' and offer.user:
        from accounts.models import Address
        firm_addresses = Address.objects.filter(user=offer.user, is_active=True)
    
    return render(
        request,
        "products/pharmacy_offer_detail.html",
        {
            "offer": offer,
            "firm_addresses": firm_addresses
        }
    )


@role_required("eczane")
@role_required("eczane")
def pharmacy_approve_offer(request, offer_id):
    """Teklifi onayla - yetki kontrolü ile"""
    offer = get_object_or_404(Offer, id=offer_id, status="sent")
    
    if request.method == "POST":
        # Yetki kontrolü
        can_approve, error_msg = can_user_approve_offer(request.user, offer)
        
        if not can_approve:
            messages.error(request, error_msg)
            return redirect("pharmacy_offer_detail", offer_id=offer.id)
        
        # Onay
        offer.status = "approved"
        offer.approved_at = timezone.now()
        offer.approved_by = request.user
        offer.rejected_at = None
        offer.reject_reason = ""
        offer.rejected_by = None
        offer.save()
        
        # Log kaydet
        log_activity(
            user=request.user,
            action='offer_approved',
            description=f"#{offer.id} nolu teklif onaylandı. Tutar: {offer.gross_total_price():,.2f} TL",
            offer=offer,
            target_user=offer.user,
            request=request
        )
        
        # Firmaya bildirim
        notify_on_offer_status_change(offer, 'sent', 'approved')
        
        messages.success(request, "Teklif onaylandı!")
        
    return redirect("pharmacy_offer_detail", offer_id=offer.id)


@role_required("eczane")
def pharmacy_reject_offer(request, offer_id):
    """Teklifi reddet - yetki kontrolü ile"""
    offer = get_object_or_404(Offer, id=offer_id, status="sent")
    
    if request.method == "POST":
        # Yetki kontrolü
        can_reject, error_msg = can_user_reject_offer(request.user, offer)
        
        if not can_reject:
            messages.error(request, error_msg)
            return redirect("pharmacy_offer_detail", offer_id=offer.id)
        
        reason = request.POST.get("reject_reason", "").strip()
        
        # Red
        offer.status = "rejected"
        offer.rejected_at = timezone.now()
        offer.rejected_by = request.user
        offer.approved_at = None
        offer.approved_by = None
        offer.reject_reason = reason
        offer.save()
        
        # Log kaydet
        log_activity(
            user=request.user,
            action='offer_rejected',
            description=f"#{offer.id} nolu teklif reddedildi. Sebep: {reason}",
            offer=offer,
            target_user=offer.user,
            request=request
        )
        
        # Firmaya bildirim
        notify_on_offer_status_change(offer, 'sent', 'rejected')
        
        messages.success(request, "Teklif reddedildi.")
        
    return redirect("pharmacy_offer_detail", offer_id=offer.id)

@role_required("eczane")
def pharmacy_update_invoice(request, offer_id):
    """Fatura No, Fatura Tarihi, Termin Tarihi kaydet - yetki kontrolü ile"""
    offer = get_object_or_404(Offer, id=offer_id)

    if offer.status != "approved":
        messages.error(request, "Fatura bilgileri sadece onaylanmış tekliflere eklenebilir.")
        return redirect("pharmacy_offer_detail", offer_id=offer.id)
    
    # Yetki kontrolü
    can_manage, error_msg = can_user_manage_invoice(request.user, offer)
    
    if not can_manage:
        messages.error(request, error_msg)
        return redirect("pharmacy_offer_detail", offer_id=offer.id)

    # Eğer fatura zaten dolu ve revize onaylanmamışsa block
    if offer.invoice_number and offer.invoice_date and offer.delivery_deadline:
        if not offer.invoice_revision_approved:
            messages.error(request, "Fatura bilgileri zaten kaydedilmiş. Revize talebi yapınız.")
            return redirect("pharmacy_offer_detail", offer_id=offer.id)

    if request.method == "POST":
        offer.invoice_number = request.POST.get("invoice_number", "").strip() or None

        invoice_date_raw = request.POST.get("invoice_date", "").strip()
        offer.invoice_date = invoice_date_raw if invoice_date_raw else None

        delivery_deadline_raw = request.POST.get("delivery_deadline", "").strip()
        offer.delivery_deadline = delivery_deadline_raw if delivery_deadline_raw else None

        # Revize sonrası kaydet → flag'leri sıfırla
        offer.invoice_revision_pending = False
        offer.invoice_revision_approved = False
        offer.invoice_revision_reason = None
        offer.invoice_revision_requested_by = None

        offer.save()
        messages.success(request, "Fatura bilgileri kaydedildi.")

    return redirect("pharmacy_offer_detail", offer_id=offer.id)

@role_required("eczane")
def pharmacy_request_invoice_revision(request, offer_id):
    """Personel → yöneticiye fatura revize talebi"""
    offer = get_object_or_404(Offer, id=offer_id)

    if offer.status != "approved":
        messages.error(request, "Geçersiz teklif durumu.")
        return redirect("pharmacy_offer_detail", offer_id=offer.id)

    if request.method == "POST":
        reason = request.POST.get("revision_reason", "").strip()
        if not reason:
            messages.error(request, "Revize nedeni zorunludur.")
            return redirect("pharmacy_offer_detail", offer_id=offer.id)

        offer.invoice_revision_pending = True
        offer.invoice_revision_approved = False
        offer.invoice_revision_reason = reason
        offer.invoice_revision_requested_by = request.user
        offer.save()
        messages.success(request, "Revize talebi yöneticiye iletildi.")

    return redirect("pharmacy_offer_detail", offer_id=offer.id)


@role_required("eczane")
def pharmacy_approve_invoice_revision(request, offer_id):
    """Yönetici → revize talebi onaylar"""
    offer = get_object_or_404(Offer, id=offer_id)

    if request.method == "POST":
        offer.invoice_revision_pending = False
        offer.invoice_revision_approved = True
        offer.save()
        messages.success(request, "Fatura revize talebi onaylandı.")

    return redirect("pharmacy_offer_detail", offer_id=offer.id)


@role_required("eczane")
def pharmacy_reject_invoice_revision(request, offer_id):
    """Yönetici → revize talebi reddeder"""
    offer = get_object_or_404(Offer, id=offer_id)

    if request.method == "POST":
        offer.invoice_revision_pending = False
        offer.invoice_revision_approved = False
        offer.invoice_revision_reason = None
        offer.invoice_revision_requested_by = None
        offer.save()
        messages.success(request, "Fatura revize talebi reddedildi.")

    return redirect("pharmacy_offer_detail", offer_id=offer.id)


@role_required("eczane")
def pharmacy_update_discounts(request, offer_id):
    """İskonto uygula - yetki kontrolü ile"""
    offer = get_object_or_404(Offer, id=offer_id)

    if offer.status != "sent":
        messages.error(request, "Bu teklif onaylandı/kapandı. Artık değişiklik yapılamaz.")
        return redirect("pharmacy_offer_detail", offer_id=offer.id)

    if request.method == "POST":
        # En yüksek iskonto oranını bul (yetki kontrolü için)
        max_discount = Decimal("0")
        
        # Item bazlı iskontoları kontrol et
        for item in offer.items.all():
            dtype = request.POST.get(f"discount_type_{item.id}", "none")
            dval_raw = request.POST.get(f"discount_value_{item.id}", "0").replace(",", ".").strip()
            
            try:
                dval = Decimal(dval_raw or "0")
            except:
                dval = Decimal("0")
            
            if dtype == "percentage" and dval > max_discount:
                max_discount = dval
        
        # Genel toplam iskontu kontrol et
        overall_dtype = request.POST.get("overall_discount_type", "none")
        overall_dval_raw = request.POST.get("overall_discount_value", "0").replace(",", ".").strip()
        
        try:
            overall_dval = Decimal(overall_dval_raw or "0")
        except:
            overall_dval = Decimal("0")
        
        if overall_dtype == "percentage" and overall_dval > max_discount:
            max_discount = overall_dval
        
        # Yetki kontrolü - en yüksek iskonto oranı için
        can_apply, error_msg = can_user_apply_discount(request.user, offer, float(max_discount))
        
        if not can_apply:
            messages.error(request, error_msg)
            return redirect("pharmacy_offer_detail", offer_id=offer.id)
        
        # --- İtem bazlı iskontoları kaydet ---
        for item in offer.items.all():
            dtype = request.POST.get(f"discount_type_{item.id}", "none")
            dval_raw = request.POST.get(f"discount_value_{item.id}", "0").replace(",", ".").strip()

            try:
                dval = Decimal(dval_raw or "0")
            except:
                dval = Decimal("0")

            note = request.POST.get(f"note_{item.id}", "").strip()

            item.discount_type = dtype
            item.discount_value = dval
            item.note = note
            item.save()

        # --- Genel toplam iskontu kaydet ---
        offer.overall_discount_type = overall_dtype
        offer.overall_discount_value = overall_dval
        offer.save()

        messages.success(request, "İskontolar kaydedildi.")
    
    return redirect("pharmacy_offer_detail", offer_id=offer.id)

# =======================
# FİRMA – TEKLİFLERİM
# =======================
@role_required("firma")
def my_offers(request):
    """Firma kullanıcısının teklifleri - Yönetici tümünü, personel sadece kendisini görür"""
    # Yönetici kendi firmasının tüm tekliflerini görebilir
    if request.user.can_view_all_offers:
    # Kendi firmasından olan kullanıcıların teklifleri
        firm_users = User.objects.filter(
            role='firma',
            manager=request.user.manager if not request.user.is_manager else None
        ) | User.objects.filter(id=request.user.id)
    
        if request.user.is_manager:
            # Yönetici ise: kendisi + bağlı personellerin teklifleri
            firm_users = User.objects.filter(manager=request.user) | User.objects.filter(id=request.user.id)
    
        offers = Offer.objects.filter(user__in=firm_users).exclude(status="draft").order_by("-created_at")
    
        # Onay bekleyen teklifler (sadece kendi firmasından)
        pending_approvals = Offer.objects.filter(
            user__in=firm_users,
            manager_approval_pending=True,
            status="draft"
        ).order_by("-created_at")

    else:
        # Personel sadece kendi tekliflerini görür
        offers = Offer.objects.filter(
            user=request.user
        ).exclude(status="draft").order_by("-created_at")
        
        # Personelin onay bekleyen teklifleri
        pending_approvals = Offer.objects.filter(
            user=request.user,
            manager_approval_pending=True,
            status="draft"
        ).order_by("-created_at")
    
    stats = {
        'total': offers.count(),
        'sent': offers.filter(status='sent').count(),
        'approved': offers.filter(status='approved').count(),
        'rejected': offers.filter(status='rejected').count(),
        'pending_approval': pending_approvals.count(),
    }
    
    return render(request, "products/my_offers_dashboard.html", {
        "offers": offers,
        "pending_approvals": pending_approvals,
        "stats": stats,
        "is_manager": request.user.is_manager,
    })


@role_required("firma")
def my_offer_detail(request, offer_id):
    """Teklif detayı - yetki kontrolü ile"""
    offer = get_object_or_404(Offer, id=offer_id)
    
    # Görüntüleme yetkisi kontrolü
    if not can_user_view_offer(request.user, offer):
        messages.error(request, "Bu teklifi görüntüleme yetkiniz yok.")
        return redirect("my_offers")

    if offer.status == "draft" and not offer.manager_approval_pending:
        raise Http404("Bu teklife erişim yok")

    return render(request, "products/my_offer_detail.html", {
        "offer": offer,
        "is_manager": request.user.is_manager,
    })


# =======================
# FİRMA – REVİZE SİSTEMİ
# =======================
@role_required("firma")
def revise_offer(request, offer_id):
    """Reddedilen teklifi revize et"""
    original_offer = get_object_or_404(Offer, id=offer_id)
    
    # Yetki kontrolü
    if not can_user_view_offer(request.user, original_offer):
        messages.error(request, "Bu teklifi görüntüleme yetkiniz yok.")
        return redirect('my_offers')
    
    # Revize yetkisi kontrolü (rejected olduğu için edit_offer çalışmaz, özel kontrol)
    can_revise = False
    
    # Admin/Superadmin her şeyi yapabilir
    if request.user.is_superuser or request.user.is_staff:
        can_revise = True
    # Yönetici her zaman revize edebilir
    elif request.user.is_manager and original_offer.user.role == request.user.role:
        can_revise = True
    # Kendi teklifi + revize yetkisi
    elif original_offer.user == request.user and request.user.can_revise_offer:
        can_revise = True
    # Tüm teklifleri düzenleyebilir yetkisi
    elif hasattr(request.user, 'can_edit_all_offers') and request.user.can_edit_all_offers and original_offer.user.role == request.user.role:
        can_revise = True
    
    if not can_revise:
        messages.error(request, "Teklifi revize etme yetkiniz yok.")
        return redirect('my_offers')
    
    if original_offer.status != 'rejected':
        messages.error(request, "Sadece reddedilen teklifler revize edilebilir.")
        return redirect('my_offers')
    
    if request.method == 'POST':
        revision_note = request.POST.get('revision_note', '').strip()
        
        if original_offer.original_offer:
            root_offer = original_offer.original_offer
        else:
            root_offer = original_offer
        
        latest_rev = root_offer.revisions.order_by('-revision_number').first()
        if latest_rev:
            latest_revision_number = latest_rev.revision_number + 1
        else:
            latest_revision_number = 2
        
        new_offer = Offer.objects.create(
            user=request.user,
            original_offer=root_offer,
            revision_number=latest_revision_number,
            revision_note=revision_note,
            revised_at=timezone.now(),
            status='draft'
        )
        
        for item in original_offer.items.all():
            OfferItem.objects.create(
                offer=new_offer,
                product=item.product,
                quantity=item.quantity,
                unit_price=item.unit_price,
                vat_rate=item.vat_rate,
                discount_type=item.discount_type,
                discount_value=item.discount_value,
                note=item.note
            )
        
        original_offer.status = 'revised'
        original_offer.save()
        
        messages.success(request, f"Teklif revize edildi! Revizyon #{latest_revision_number} oluşturuldu.")
        return redirect('offer')
    
    return render(request, 'products/revise_offer.html', {
        'original_offer': original_offer
    })


@role_required("firma")
def delete_offer_item(request, item_id):
    item = get_object_or_404(OfferItem, id=item_id)
    
    if item.offer.user != request.user or item.offer.status != 'draft':
        raise PermissionDenied
    
    item.delete()
    messages.success(request, "Ürün sepetten kaldırıldı.")
    return redirect('offer')


@role_required("firma")
def update_offer_item(request, item_id):
    item = get_object_or_404(OfferItem, id=item_id)
    
    if item.offer.user != request.user or item.offer.status != 'draft':
        raise PermissionDenied
    
    if request.method == 'POST':
        quantity = request.POST.get('quantity', 1)
        try:
            quantity = int(quantity)
            if quantity < 1:
                quantity = 1
        except:
            quantity = 1
        
        item.quantity = quantity
        item.save()
        messages.success(request, "Ürün miktarı güncellendi.")
    
    return redirect('offer')


@role_required("firma")
def view_offer_history(request, offer_id):
    """Teklif geçmişini görüntüle"""
    
    # Teklifi bul
    offer = get_object_or_404(Offer, id=offer_id)
    
    # Yetki kontrolü
    if request.user.is_superuser or request.user.is_staff:
        # Admin her şeyi görebilir
        pass
    elif request.user.is_manager:
        # Yönetici kendi firmasının/eczanesinin tekliflerini görebilir
        if offer.user.role != request.user.role:
            messages.error(request, "Bu teklifi görüntüleme yetkiniz yok.")
            return redirect('my_offers')
    elif hasattr(request.user, 'can_view_all_offers') and request.user.can_view_all_offers:
        # Personel "tüm teklifleri görebilir" yetkisi varsa
        if offer.user.role != request.user.role:
            messages.error(request, "Bu teklifi görüntüleme yetkiniz yok.")
            return redirect('my_offers')
    else:
        # Normal personel sadece kendi tekliflerini görebilir
        if offer.user != request.user:
            messages.error(request, "Bu teklifi görüntüleme yetkiniz yok.")
            return redirect('my_offers')
    
    all_revisions = offer.get_all_revisions()
    
    return render(request, 'products/offer_history.html', {
        'offer': offer,
        'all_revisions': all_revisions
    })


# =======================
# ECZANE – DASHBOARD
# =======================
@role_required("eczane")
def pharmacy_dashboard(request):
    sent_offers = Offer.objects.filter(status='sent')
    
    revised_original_ids = sent_offers.filter(
        original_offer__isnull=False
    ).values_list('original_offer_id', flat=True).distinct()
    
    latest_sent_offers = sent_offers.exclude(
        id__in=revised_original_ids
    )
    
    new_offers_count = latest_sent_offers.count()
    
    stats = {
        'pending': new_offers_count,
        'approved': Offer.objects.filter(status='approved').count(),
        'rejected': Offer.objects.filter(status='rejected').count(),
        'total': Offer.objects.exclude(status='draft').count(),
    }
    
    recent_offers = latest_sent_offers.order_by('-sent_at')[:5]
    
    return render(request, 'products/pharmacy_dashboard.html', {
        'new_offers_count': new_offers_count,
        'stats': stats,
        'recent_offers': recent_offers,
    })


@role_required("eczane")
def pharmacy_offers_by_status(request, status):
    valid_statuses = ['sent', 'approved', 'rejected']
    
    if status not in valid_statuses:
        messages.error(request, "Geçersiz durum!")
        return redirect('pharmacy_dashboard')
    
    offers = Offer.objects.filter(status=status).order_by('-sent_at')
    
    status_names = {
        'sent': 'Bekleyen',
        'approved': 'Onaylanan',
        'rejected': 'Reddedilen'
    }
    
    return render(request, 'products/pharmacy_offers_filtered.html', {
        'offers': offers,
        'status': status,
        'status_name': status_names[status],
        'count': offers.count()
    })


# =======================
# FİRMA – DURUM FILTRELEME
# =======================
@role_required("firma")
def my_offers_by_status(request, status):
    valid_statuses = ['sent', 'approved', 'rejected', 'revised']
    
    if status not in valid_statuses:
        messages.error(request, "Geçersiz durum!")
        return redirect('my_offers')
    
    offers = Offer.objects.filter(
        user=request.user,
        status=status
    ).order_by('-created_at')
    
    status_names = {
        'sent': 'Gönderilen',
        'approved': 'Onaylanan',
        'rejected': 'Reddedilen',
        'revised': 'Revize Edilen'
    }
    
    return render(request, 'products/my_offers_filtered.html', {
        'offers': offers,
        'status': status,
        'status_name': status_names[status],
        'count': offers.count()
    })


# =======================
# ADMİN DASHBOARD (tek kez)
# =======================
@login_required
def admin_dashboard(request):
    if not request.user.is_superuser:
        if hasattr(request.user, 'role') and request.user.role:
            if request.user.role == 'eczane':
                return redirect('pharmacy_dashboard')
            elif request.user.role == 'firma':
                return redirect('my_offers')
        return redirect('product_list')
    
    # Kullanıcı istatistikleri
    total_users = User.objects.count()
    firma_users = User.objects.filter(role='firma').count()
    eczane_users = User.objects.filter(role='eczane').count()
    pending_approvals = User.objects.filter(is_approved=False).count()
    
    # Firma yöneticileri ve personelleri
    firma_managers = User.objects.filter(role='firma', is_manager=True).count()
    firma_staff = User.objects.filter(role='firma', is_manager=False).count()
    
    # Eczane eczacıları ve personelleri
    eczane_managers = User.objects.filter(role='eczane', is_manager=True).count()
    eczane_staff = User.objects.filter(role='eczane', is_manager=False).count()
    
    # Teklif istatistikleri
    total_offers = Offer.objects.count()
    sent_offers = Offer.objects.filter(status='sent').count()
    approved_offers = Offer.objects.filter(status='approved').count()
    rejected_offers = Offer.objects.filter(status='rejected').count()
    
    # Yüksek tutarlı teklifler - KALDIRILDI (gross_total_price field yok)
    high_value_offers = 0  # Şimdilik 0
    
    # Bekleyen onaylar (firma tarafı)
    pending_manager_approvals = Offer.objects.filter(manager_approval_pending=True).count()
    
    # Son aktiviteler
    recent_users = User.objects.order_by('-date_joined')[:10]
    recent_offers = Offer.objects.order_by('-created_at')[:10]
    
    return render(request, 'admin/admin_dashboard.html', {
        'total_users': total_users,
        'firma_users': firma_users,
        'eczane_users': eczane_users,
        'pending_approvals': pending_approvals,
        'firma_managers': firma_managers,
        'firma_staff': firma_staff,
        'eczane_managers': eczane_managers,
        'eczane_staff': eczane_staff,
        'total_offers': total_offers,
        'sent_offers': sent_offers,
        'approved_offers': approved_offers,
        'rejected_offers': rejected_offers,
        'high_value_offers': high_value_offers,
        'pending_manager_approvals': pending_manager_approvals,
        'recent_users': recent_users,
        'recent_offers': recent_offers,
    })

# =======================
# KULLANICI YÖNETİMİ
# =======================
def admin_users(request):
    """Modern kullanıcı yönetim sayfası"""
    if not request.user.is_superuser:
        return redirect('admin_dashboard')
    
    # Filtreleme
    role_filter = request.GET.get('role', 'all')
    status_filter = request.GET.get('status', 'all')
    search_query = request.GET.get('search', '').strip()
    
    users = User.objects.all().order_by('-date_joined')
    
    # Role filtresi
    if role_filter != 'all':
        users = users.filter(role=role_filter)
    
    # Status filtresi
    if status_filter == 'approved':
        users = users.filter(is_approved=True)
    elif status_filter == 'pending':
        users = users.filter(is_approved=False)
    
    # Arama
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )
    
    return render(request, 'admin/users_management.html', {
        'users': users,
        'role_filter': role_filter,
        'status_filter': status_filter,
        'search_query': search_query,
    })


def admin_user_approve(request, user_id):
    """Kullanıcıyı onayla"""
    if not request.user.is_superuser:
        return redirect('admin_dashboard')
    
    user = get_object_or_404(User, id=user_id)
    user.is_approved = True
    user.save()
    messages.success(request, f"{user.username} başarıyla onaylandı.")
    return redirect(request.META.get('HTTP_REFERER', 'admin_users'))


def admin_user_reject(request, user_id):
    """Kullanıcı onayını iptal et"""
    if not request.user.is_superuser:
        return redirect('admin_dashboard')
    
    user = get_object_or_404(User, id=user_id)
    user.is_approved = False
    user.save()
    messages.warning(request, f"{user.username} onayı iptal edildi.")
    return redirect(request.META.get('HTTP_REFERER', 'admin_users'))


def admin_user_delete(request, user_id):
    """Kullanıcıyı sil"""
    if not request.user.is_superuser:
        return redirect('admin_dashboard')
    
    user = get_object_or_404(User, id=user_id)
    if user.is_superuser:
        messages.error(request, "Superuser silinemez!")
        return redirect('admin_users')
    
    username = user.username
    user.delete()
    messages.success(request, f"{username} silindi.")
    return redirect('admin_users')


def admin_user_edit(request, user_id):
    """Kullanıcı bilgilerini düzenle"""
    if not request.user.is_superuser:
        return redirect('admin_dashboard')
    
    user = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        user.first_name = request.POST.get('first_name', '').strip()
        user.last_name = request.POST.get('last_name', '').strip()
        user.email = request.POST.get('email', '').strip()
        user.role = request.POST.get('role', user.role)
        user.is_approved = request.POST.get('is_approved') == 'on'
        
        # Yetki Yönetimi
        user.is_manager = request.POST.get('is_manager') == 'on'
        user.is_active_user = request.POST.get('is_active_user') == 'on'
        
        # Firma bilgileri
        if user.role == 'firma':
            user.company_name = request.POST.get('company_name', '').strip()
            user.company_tax_number = request.POST.get('company_tax_number', '').strip()
            user.company_phone = request.POST.get('company_phone', '').strip()
            user.company_address = request.POST.get('company_address', '').strip()
            user.company_responsible_person = request.POST.get('company_responsible_person', '').strip()
            
            # Firma Yetkileri
            user.can_create_offer = request.POST.get('can_create_offer') == 'on'
            user.can_edit_own_offer = request.POST.get('can_edit_own_offer') == 'on'
            user.can_edit_all_offers = request.POST.get('can_edit_all_offers') == 'on'
            user.can_send_offer = request.POST.get('can_send_offer') == 'on'
            user.can_view_all_offers = request.POST.get('can_view_all_offers') == 'on'
            user.can_approve_high_value_offers = request.POST.get('can_approve_high_value_offers') == 'on'
            user.can_invite_staff = request.POST.get('can_invite_staff') == 'on'
            user.can_manage_staff_permissions = request.POST.get('can_manage_staff_permissions') == 'on'
            user.can_view_reports = request.POST.get('can_view_reports') == 'on'
            user.can_view_financial_data = request.POST.get('can_view_financial_data') == 'on'
        
        # Eczane bilgileri
        elif user.role == 'eczane':
            user.pharmacy_name = request.POST.get('pharmacy_name', '').strip()
            user.pharmacist_name = request.POST.get('pharmacist_name', '').strip()
            user.pharmacy_tax_number = request.POST.get('pharmacy_tax_number', '').strip()
            user.pharmacy_phone = request.POST.get('pharmacy_phone', '').strip()
            user.pharmacy_address = request.POST.get('pharmacy_address', '').strip()
            user.pharmacy_license_number = request.POST.get('pharmacy_license_number', '').strip()
            
            # Eczane Yetkileri
            user.can_approve_pharmacy_offers = request.POST.get('can_approve_pharmacy_offers') == 'on'
            user.can_reject_pharmacy_offers = request.POST.get('can_reject_pharmacy_offers') == 'on'
            user.can_approve_high_value_pharmacy_offers = request.POST.get('can_approve_high_value_pharmacy_offers') == 'on'
            user.can_apply_discount = request.POST.get('can_apply_discount') == 'on'
            user.can_apply_high_discount = request.POST.get('can_apply_high_discount') == 'on'
            user.can_enter_invoice = request.POST.get('can_enter_invoice') == 'on'
            user.can_edit_invoice = request.POST.get('can_edit_invoice') == 'on'
            user.can_approve_pharmacy_staff = request.POST.get('can_approve_pharmacy_staff') == 'on'
            user.can_manage_pharmacy_staff = request.POST.get('can_manage_pharmacy_staff') == 'on'
        
        user.save()
        messages.success(request, f"{user.username} bilgileri ve yetkileri güncellendi.")
        return redirect('admin_users')
    
    return render(request, 'admin/user_edit.html', {'edit_user': user})

# =======================
# FAVORİ SİSTEMİ
# =======================
@role_required("firma")
def toggle_favorite_product(request, product_id):
    """Ürünü favorilere ekle/çıkar"""
    product = get_object_or_404(Product, id=product_id)
    
    from products.models import FavoriteProduct
    favorite = FavoriteProduct.objects.filter(user=request.user, product=product).first()
    
    if favorite:
        favorite.delete()
        messages.success(request, f"{product.name} favorilerden çıkarıldı.")
    else:
        FavoriteProduct.objects.create(user=request.user, product=product)
        messages.success(request, f"{product.name} favorilere eklendi!")
    
    return redirect(request.META.get('HTTP_REFERER', 'product_list'))


@role_required("firma")
def favorite_products(request):
    """Favori ürünler listesi"""
    from products.models import FavoriteProduct
    favorites = FavoriteProduct.objects.filter(user=request.user).select_related('product')
    
    return render(request, 'products/favorite_products.html', {
        'favorites': favorites
    })


@role_required("firma")
def toggle_favorite_draft(request, offer_id):
    """Teklif taslağını favorilere ekle/çıkar"""
    offer = get_object_or_404(Offer, id=offer_id, user=request.user, status='draft')
    
    from products.models import FavoriteDraft
    favorite = FavoriteDraft.objects.filter(user=request.user, offer=offer).first()
    
    if favorite:
        favorite.delete()
        messages.success(request, "Taslak favorilerden çıkarıldı.")
    else:
        note = request.POST.get('note', '').strip() if request.method == 'POST' else ''
        FavoriteDraft.objects.create(user=request.user, offer=offer, note=note)
        messages.success(request, "Taslak favorilere eklendi!")
    
    return redirect(request.META.get('HTTP_REFERER', 'offer'))


@role_required("firma")
def favorite_drafts(request):
    """Favori taslaklar listesi"""
    from products.models import FavoriteDraft
    favorites = FavoriteDraft.objects.filter(user=request.user).select_related('offer')
    
    return render(request, 'products/favorite_drafts.html', {
        'favorites': favorites
    })

# =======================
# BİLDİRİM SİSTEMİ
# =======================
@login_required
def get_notifications(request):
    """Kullanıcının bildirimlerini getir"""
    notifications = Notification.objects.filter(
        user=request.user
    ).order_by('-created_at')[:20]
    
    unread_count = notifications.filter(is_read=False).count()
    
    return {
        'notifications': notifications,
        'unread_count': unread_count
    }


@login_required
@require_POST
def mark_notification_read(request, notification_id):
    """Bildirimi okundu olarak işaretle"""
    notification = get_object_or_404(Notification, id=notification_id, user=request.user)
    
    if not notification.is_read:
        notification.is_read = True
        notification.read_at = timezone.now()
        notification.save()
    
    return JsonResponse({'status': 'success'})


@login_required
@require_POST
def mark_all_notifications_read(request):
    """Tüm bildirimleri okundu olarak işaretle"""
    Notification.objects.filter(
        user=request.user,
        is_read=False
    ).update(
        is_read=True,
        read_at=timezone.now()
    )
    
    return JsonResponse({'status': 'success'})

"""
Bu kodları products/views.py dosyasındaki mevcut helper fonksiyonların ALTINA ekleyin
"""

def can_user_approve_offer(user, offer):
    """Kullanıcı bu teklifi onaylayabilir mi? (Eczane tarafı)"""
    
    # Admin/Superadmin her şeyi onaylayabilir
    if user.is_superuser or user.is_staff:
        return True, None
    
    # Sadece eczane kullanıcıları onaylayabilir
    if user.role != 'eczane':
        return False, "Sadece eczane kullanıcıları teklif onaylayabilir."
    
    # Teklif gönderilmiş olmalı
    if offer.status != 'sent':
        return False, "Sadece gönderilmiş teklifler onaylanabilir."
    
    # Eczacı (yönetici) her şeyi onaylayabilir
    if user.is_manager:
        return True, None
    
    # Personel için yetki kontrolü
    if not hasattr(user, 'can_approve_pharmacy_offers'):
        return False, "Teklif onaylama yetkiniz yok."
    
    if not user.can_approve_pharmacy_offers:
        return False, "Teklif onaylama yetkiniz yok."
    
    # 50K+ teklifler için özel yetki gerekli
    from decimal import Decimal
    if hasattr(offer, 'gross_total_price') and offer.gross_total_price() >= Decimal('50000'):
        if not (hasattr(user, 'can_approve_high_value_pharmacy_offers') and user.can_approve_high_value_pharmacy_offers):
            return False, "50.000 TL üzeri teklifleri onaylama yetkiniz yok. Eczacınızla iletişime geçin."
    
    return True, None


def can_user_reject_offer(user, offer):
    """Kullanıcı bu teklifi reddedebilir mi? (Eczane tarafı)"""
    
    # Admin/Superadmin her şeyi reddedebilir
    if user.is_superuser or user.is_staff:
        return True, None
    
    # Sadece eczane kullanıcıları reddedebilir
    if user.role != 'eczane':
        return False, "Sadece eczane kullanıcıları teklif reddedebilir."
    
    # Teklif gönderilmiş olmalı
    if offer.status != 'sent':
        return False, "Sadece gönderilmiş teklifler reddedilebilir."
    
    # Eczacı (yönetici) her şeyi reddedebilir
    if user.is_manager:
        return True, None
    
    # Personel için yetki kontrolü
    if not hasattr(user, 'can_reject_pharmacy_offers'):
        return False, "Teklif reddetme yetkiniz yok."
    
    if not user.can_reject_pharmacy_offers:
        return False, "Teklif reddetme yetkiniz yok."
    
    return True, None


def can_user_apply_discount(user, offer, discount_percent):
    """Kullanıcı iskonto uygulayabilir mi? (Eczane tarafı)"""
    
    # Admin/Superadmin her şeyi yapabilir
    if user.is_superuser or user.is_staff:
        return True, None
    
    # Sadece eczane kullanıcıları iskonto uygulayabilir
    if user.role != 'eczane':
        return False, "Sadece eczane kullanıcıları iskonto uygulayabilir."
    
    # Eczacı (yönetici) her şeyi yapabilir
    if user.is_manager:
        return True, None
    
    # Personel için yetki kontrolü
    if not hasattr(user, 'can_apply_discount'):
        return False, "İskonto uygulama yetkiniz yok."
    
    if not user.can_apply_discount:
        return False, "İskonto uygulama yetkiniz yok."
    
    # %20+ iskonto için özel yetki gerekli
    if discount_percent >= 20:
        if not (hasattr(user, 'can_apply_high_discount') and user.can_apply_high_discount):
            return False, "%20 ve üzeri iskonto uygulama yetkiniz yok. Eczacınızla iletişime geçin."
    
    return True, None


def can_user_manage_invoice(user, offer):
    """Kullanıcı fatura işlemleri yapabilir mi? (Eczane tarafı)"""
    
    # Admin/Superadmin her şeyi yapabilir
    if user.is_superuser or user.is_staff:
        return True, None
    
    # Sadece eczane kullanıcıları fatura yönetebilir
    if user.role != 'eczane':
        return False, "Sadece eczane kullanıcıları fatura yönetebilir."
    
    # Sadece onaylı teklifler için fatura girilebilir
    if offer.status != 'approved':
        return False, "Sadece onaylı teklifler için fatura girilebilir."
    
    # Eczacı (yönetici) her şeyi yapabilir
    if user.is_manager:
        return True, None
    
    # Personel için yetki kontrolü
    if not hasattr(user, 'can_enter_invoice'):
        return False, "Fatura girme yetkiniz yok."
    
    if not user.can_enter_invoice:
        return False, "Fatura girme yetkiniz yok."
    
    return True, None

"""
Bu kodu products/views.py dosyasının SONUNA ekleyin
"""

@login_required
def assign_delivery_addresses(request, offer_id):
    """Onaylanan teklif için ürünlere teslimat adresi ata"""
    offer = get_object_or_404(Offer, id=offer_id)
    
    # Sadece firma kullanıcısı kendi teklifine adres atayabilir
    if offer.user != request.user:
        messages.error(request, "Bu teklif size ait değil.")
        return redirect('my_offers')
    
    # Sadece onaylanmış tekliflere adres atanabilir
    if offer.status != 'approved':
        messages.error(request, "Sadece onaylanmış tekliflere teslimat adresi atanabilir.")
        return redirect('my_offer_detail', offer_id=offer.id)
    
    from accounts.models import Address
    firm_addresses = Address.objects.filter(user=request.user, is_active=True)
    
    if request.method == 'POST':
        # Her ürün için seçilen adresi kaydet
        for item in offer.items.all():
            address_id = request.POST.get(f'address_{item.id}')
            if address_id:
                try:
                    address = Address.objects.get(id=address_id, user=request.user)
                    item.delivery_address = address
                    item.save()
                except Address.DoesNotExist:
                    pass
        
        messages.success(request, "Teslimat adresleri kaydedildi.")
        return redirect('my_offer_detail', offer_id=offer.id)
    
    return render(request, 'products/assign_delivery_addresses.html', {
        'offer': offer,
        'firm_addresses': firm_addresses
    })

@login_required
def test_navbar(request):
    """Navbar test sayfası"""
    return render(request, 'products/test_navbar.html')

@login_required
def pharmacy_product_management(request):
    """Eczane ürün yönetimi - sadece eczacı veya yetkili personel"""
    
    # Admin veya eczane kullanıcıları erişebilir
    if not (request.user.is_superuser or request.user.is_staff or request.user.role == 'eczane'):
        messages.error(request, "Bu sayfaya erişim yetkiniz yok.")
        return redirect('redirect_after_login')
    
    # Yetki kontrolü
    if request.user.role == 'eczane':  # Admin için yetki kontrolü yapma
        if not request.user.is_manager and not request.user.can_manage_products:
            messages.error(request, "Ürün yönetimi yetkiniz yok.")
            return redirect('pharmacy_dashboard')
    
    # Tüm ürünleri getir
    products_list = Product.objects.all().order_by('-id')

    # Arama filtresi
    search_query = request.GET.get('search', '').strip()
    if search_query:
        products_list = products_list.filter(
            Q(name__icontains=search_query) | 
            Q(barcode__icontains=search_query)
        )
    
    # Sayfalama
    paginator = Paginator(products_list, 50)
    page_number = request.GET.get('page', 1)
    products = paginator.get_page(page_number)
    
    return render(request, 'products/pharmacy_product_management.html', {
        'products': products,
        'can_add': True,
        'can_delete': request.user.is_superuser or request.user.is_staff or request.user.is_manager,
    })


@login_required
def pharmacy_add_product(request):
    """Tek ürün ekleme"""
    
    # Yetki kontrolü
    if not (request.user.is_superuser or request.user.is_staff or request.user.role == 'eczane'):
        messages.error(request, "Bu sayfaya erişim yetkiniz yok.")
        return redirect('redirect_after_login')
    
    
    if not request.user.is_manager and not request.user.can_manage_products:
        messages.error(request, "Ürün ekleme yetkiniz yok.")
        return redirect('pharmacy_dashboard')
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip().upper()
        barcode = request.POST.get('barcode', '').strip()
        unit_price_str = request.POST.get('unit_price', '0').replace(',', '.')
        vat_rate_str = request.POST.get('vat_rate', '10').replace(',', '.')
        
        try:
            price = Decimal(unit_price_str)
            vat_rate = int(float(vat_rate_str))
        except:
            messages.error(request, "Geçersiz fiyat veya KDV oranı.")
            return render(request, 'products/pharmacy_add_product.html')
        
        if not name:
            messages.error(request, "Ürün adı zorunludur.")
            return render(request, 'products/pharmacy_add_product.html')
        
        # Ürün oluştur
        Product.objects.create(
            name=name,
            barcode=barcode,
            price=price,
            vat_rate=vat_rate
        )
        
        messages.success(request, f"{name} ürünü eklendi.")
        return redirect('pharmacy_product_management')
    
    return render(request, 'products/pharmacy_add_product.html')


@login_required
def pharmacy_import_products_excel(request):
    """Excel ile toplu ürün ekleme"""
    
    # Yetki kontrolü
    if request.user.role != 'eczane':
        messages.error(request, "Bu sayfaya sadece eczane kullanıcıları erişebilir.")
        return redirect('redirect_after_login')
    
    if not request.user.is_manager and not request.user.can_manage_products:
        messages.error(request, "Ürün ekleme yetkiniz yok.")
        return redirect('pharmacy_dashboard')
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        try:
            # Excel dosyasını oku
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            
            added_count = 0
            error_count = 0
            
            # Satırları oku (başlık satırını atla)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Ürün adı boşsa atla
                    continue
                
                try:
                    name = str(row[0]).strip().upper()
                    barcode = str(row[1]).strip() if row[1] else ''
                    price = Decimal(str(row[2]).replace(',', '.')) if row[2] else Decimal('0')
                    vat_rate = int(float(str(row[3]).replace(',', '.'))) if row[3] else 10

                    Product.objects.create(
                        name=name,
                        barcode=barcode,
                        price=price,
                        vat_rate=vat_rate
                    )
                    added_count += 1
                except Exception as e:
                    error_count += 1
                    continue
            
            if added_count > 0:
                messages.success(request, f"{added_count} ürün başarıyla eklendi.")
            if error_count > 0:
                messages.warning(request, f"{error_count} ürün eklenirken hata oluştu.")
            
            return redirect('pharmacy_product_management')
            
        except Exception as e:
            messages.error(request, f"Excel dosyası işlenirken hata: {str(e)}")
    
    return render(request, 'products/pharmacy_import_excel.html')


@login_required
def pharmacy_delete_product(request, product_id):
    """Ürün silme - sadece admin"""
    
    # Admin veya eczacı silebilir
    if not (request.user.is_superuser or request.user.is_staff or request.user.is_manager):
        messages.error(request, "Ürün silme yetkiniz yok.")
        return redirect('pharmacy_product_management')
    
    product = get_object_or_404(Product, id=product_id)
    
    if request.method == 'POST':
        name = product.name
        product.delete()
        messages.success(request, f"{name} ürünü silindi.")
    
    return redirect('pharmacy_product_management')

@login_required
def pharmacy_update_product_price(request, product_id):
    """Eczane ürün fiyatını günceller"""
    
    # Sadece eczane kullanıcıları
    if request.user.role != 'eczane':
        messages.error(request, "Bu işlem için yetkiniz yok.")
        return redirect('pharmacy_product_management')
    
    # Yetki kontrolü: Eczacı veya yetkili personel
    if not request.user.is_manager and not request.user.can_update_product_prices:
        messages.error(request, "Fiyat güncelleme yetkiniz yok.")
        return redirect('pharmacy_product_management')
    
    product = get_object_or_404(Product, id=product_id)
    
    if request.method == 'POST':
        new_price = request.POST.get('price', '').replace(',', '.')
        
        try:
            new_price = Decimal(new_price)
            if new_price < 0:
                raise ValueError("Fiyat negatif olamaz")
            
            product.price = new_price
            product.save()  # price_updated_at otomatik güncellenir (auto_now=True)
            
            messages.success(request, f"{product.name} fiyatı güncellendi: {new_price} ₺")
        except (ValueError, InvalidOperation):
            messages.error(request, "Geçersiz fiyat değeri.")
    
    return redirect('pharmacy_product_management')


@login_required
def pharmacy_edit_product(request, product_id):
    """Eczane ürün düzenler (ad, fiyat, KDV)"""
    
    # Sadece eczane kullanıcıları
    if request.user.role != 'eczane':
        messages.error(request, "Bu işlem için yetkiniz yok.")
        return redirect('pharmacy_product_management')
    
    # Yetki kontrolü
    if not request.user.is_manager and not request.user.can_update_product_prices:
        messages.error(request, "Ürün düzenleme yetkiniz yok.")
        return redirect('pharmacy_product_management')
    
    product = get_object_or_404(Product, id=product_id)
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip().upper()
        price = request.POST.get('price', '').replace(',', '.')
        vat_rate = request.POST.get('vat_rate', '10')
        
        try:
            price = Decimal(price)
            vat_rate = int(vat_rate)
            
            if not name:
                raise ValueError("Ürün adı boş olamaz")
            
            product.name = name
            product.price = price
            product.vat_rate = vat_rate
            product.save()
            
            messages.success(request, f"{name} ürünü güncellendi.")
            return redirect('pharmacy_product_management')
        except (ValueError, InvalidOperation) as e:
            messages.error(request, f"Hata: {str(e)}")
    
    return render(request, 'products/pharmacy_edit_product.html', {'product': product})

#Dışa Aktarma -------------------------------

@login_required
def export_offers_excel(request):
    """Firma - Teklifleri Excel'e aktar"""
    
    if request.user.role != 'firma':
        messages.error(request, "Bu işlem için yetkiniz yok.")
        return redirect('my_offers')
    
    if not request.user.can_export_data and not request.user.is_manager:
        messages.error(request, "Dışa aktarma yetkiniz yok.")
        return redirect('my_offers')
    
    # Teklifleri getir
    if request.user.is_manager:
        firm_users = User.objects.filter(manager=request.user) | User.objects.filter(id=request.user.id)
    else:
        firm_users = User.objects.filter(id=request.user.id)
    
    offers = Offer.objects.filter(user__in=firm_users).exclude(status='draft').order_by('-created_at')
    
    wb = Workbook()
    
    # ========================
    # SAYFA 1: TEKLİF ÖZETİ
    # ========================
    ws1 = wb.active
    ws1.title = "Teklif Özeti"
    
    # Başlık stilleri
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', start_color='2F6FED')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Firma bilgileri başlığı
    ws1.merge_cells('A1:K1')
    ws1['A1'] = f"TEKLİF RAPORU - {request.user.company_name or request.user.get_full_name()}"
    ws1['A1'].font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    ws1['A1'].fill = PatternFill('solid', start_color='1A3A6B')
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 35
    
    ws1.merge_cells('A2:K2')
    ws1['A2'] = f"Dışa Aktarma Tarihi: {timezone.now().strftime('%d.%m.%Y %H:%M')}"
    ws1['A2'].font = Font(name='Arial', size=10, color='666666')
    ws1['A2'].alignment = Alignment(horizontal='center')
    ws1.row_dimensions[2].height = 20
    
    # Boş satır
    ws1.row_dimensions[3].height = 10
    
    # Başlıklar
    headers = ['Teklif No', 'Revizyon', 'Personel', 'Durum', 'Gönderilme Tarihi', 
               'Onay/Red Tarihi', 'Ürün Sayısı', 'KDV Hariç Toplam', 
               'KDV Toplamı', 'KDV Dahil Toplam', 'Red Nedeni']
    
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws1.row_dimensions[4].height = 30
    
    # Veri satırları
    status_map = {'sent': 'Bekliyor', 'approved': 'Onaylandı', 'rejected': 'Reddedildi', 'revised': 'Revize Edildi'}
    
    for row_idx, offer in enumerate(offers, 5):
        row_fill = PatternFill('solid', start_color='F8F9FA') if row_idx % 2 == 0 else PatternFill('solid', start_color='FFFFFF')
        
        data = [
            f"#{offer.original_offer.id if offer.original_offer else offer.id}",
            f"Rev.{offer.revision_number}" if offer.revision_number > 1 else "-",
            offer.user.get_full_name() or offer.user.username,
            status_map.get(offer.status, offer.status),
            offer.sent_at.strftime('%d.%m.%Y %H:%M') if offer.sent_at else '-',
            (offer.approved_at or offer.rejected_at).strftime('%d.%m.%Y %H:%M') if (offer.approved_at or offer.rejected_at) else '-',
            offer.items.count(),
            float(offer.items_net_after_item_discounts()),
            float(offer.items_vat_after_item_discounts()),
            float(offer.final_total),
            offer.reject_reason or '-',
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws1.cell(row=row_idx, column=col, value=value)
            cell.font = Font(name='Arial', size=10)
            cell.fill = row_fill
            cell.alignment = center_align
            
            # Durum renklendirme
            if col == 4:
                if value == 'Onaylandı':
                    cell.font = Font(name='Arial', size=10, bold=True, color='155724')
                elif value == 'Reddedildi':
                    cell.font = Font(name='Arial', size=10, bold=True, color='721C24')
                elif value == 'Bekliyor':
                    cell.font = Font(name='Arial', size=10, bold=True, color='856404')
            
            # Para formatı
            if col in [8, 9, 10]:
                cell.number_format = '#,##0.00 ₺'
    
    # Toplam satırı
    total_row = len(offers) + 5
    ws1.cell(row=total_row, column=7, value=f'=SUM(G5:G{total_row-1})').font = Font(name='Arial', bold=True)
    ws1.cell(row=total_row, column=8, value=f'=SUM(H5:H{total_row-1})').number_format = '#,##0.00 ₺'
    ws1.cell(row=total_row, column=9, value=f'=SUM(I5:I{total_row-1})').number_format = '#,##0.00 ₺'
    ws1.cell(row=total_row, column=10, value=f'=SUM(J5:J{total_row-1})').number_format = '#,##0.00 ₺'
    for col in range(1, 12):
        cell = ws1.cell(row=total_row, column=col)
        cell.fill = PatternFill('solid', start_color='2F6FED')
        cell.font = Font(name='Arial', bold=True, color='FFFFFF')
        cell.alignment = center_align
    ws1.cell(row=total_row, column=1, value='TOPLAM').font = Font(name='Arial', bold=True, color='FFFFFF')
    
    # Sütun genişlikleri
    col_widths = [12, 10, 20, 12, 18, 18, 12, 18, 15, 18, 30]
    for i, width in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = width
    
    # ========================
    # SAYFA 2: ÜRÜN DETAYLARI
    # ========================
    ws2 = wb.create_sheet("Ürün Detayları")
    
    headers2 = ['Teklif No', 'Personel', 'Durum', 'Tarih', 'Ürün Adı', 'Barkod',
                'Adet', 'Birim Fiyat (KDV Hariç)', 'KDV %', 'İskonto Tipi', 
                'İskonto Değeri', 'İskonto Tutarı', 'KDV Hariç Toplam', 
                'KDV Tutarı', 'KDV Dahil Toplam']
    
    ws2.merge_cells(f'A1:{get_column_letter(len(headers2))}1')
    ws2['A1'] = f"ÜRÜN DETAY RAPORU - {request.user.company_name or request.user.get_full_name()}"
    ws2['A1'].font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    ws2['A1'].fill = PatternFill('solid', start_color='1A3A6B')
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 35
    
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws2.row_dimensions[2].height = 30
    
    row_idx = 3
    for offer in offers:
        for item in offer.items.all():
            row_fill = PatternFill('solid', start_color='F8F9FA') if row_idx % 2 == 0 else PatternFill('solid', start_color='FFFFFF')
            
            data = [
                f"#{offer.original_offer.id if offer.original_offer else offer.id}",
                offer.user.get_full_name() or offer.user.username,
                status_map.get(offer.status, offer.status),
                offer.sent_at.strftime('%d.%m.%Y') if offer.sent_at else '-',
                item.product.name,
                item.product.barcode or '-',
                item.quantity,
                float(item.unit_price),
                item.vat_rate,
                {'none': 'Yok', 'percent': '%', 'amount': '₺'}.get(item.discount_type, '-'),
                float(item.discount_value),
                float(item.discount_amount),
                float(item.line_subtotal),
                float(item.vat_amount),
                float(item.total_price),
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws2.cell(row=row_idx, column=col, value=value)
                cell.font = Font(name='Arial', size=10)
                cell.fill = row_fill
                cell.alignment = center_align
                if col in [8, 12, 13, 14, 15]:
                    cell.number_format = '#,##0.00 ₺'
            
            row_idx += 1
    
    col_widths2 = [12, 20, 12, 12, 30, 15, 8, 20, 8, 12, 12, 15, 18, 12, 18]
    for i, width in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = width
    
    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="teklifler_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_products_excel(request):
    """Ürün listesini Excel'e aktar"""
    
    if not (request.user.is_superuser or request.user.is_staff or 
            request.user.role == 'eczane' or request.user.role == 'firma'):
        messages.error(request, "Bu işlem için yetkiniz yok.")
        return redirect('pharmacy_dashboard')
    
    products = Product.objects.all().order_by('name')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Ürün Listesi"
    
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', start_color='2F6FED')
    header_align = Alignment(horizontal='center', vertical='center')
    center_align = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A1:G1')
    ws['A1'] = "ÜRÜN LİSTESİ"
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', start_color='1A3A6B')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Dışa Aktarma Tarihi: {timezone.now().strftime('%d.%m.%Y %H:%M')} | Toplam: {products.count()} ürün"
    ws['A2'].font = Font(name='Arial', size=10, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')
    
    headers = ['Sıra', 'Ürün Adı', 'Barkod', 'Birim Fiyat (KDV Dahil)', 'KDV %', 
               'Fiyat Güncelleme Tarihi', 'Eklenme Tarihi']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws.row_dimensions[3].height = 30
    
    for row_idx, product in enumerate(products, 4):
        row_fill = PatternFill('solid', start_color='F8F9FA') if row_idx % 2 == 0 else PatternFill('solid', start_color='FFFFFF')
        
        data = [
            row_idx - 3,
            product.name,
            product.barcode or '-',
            float(product.price),
            product.vat_rate,
            product.price_updated_at.strftime('%d.%m.%Y %H:%M') if product.price_updated_at else '-',
            product.created_at.strftime('%d.%m.%Y %H:%M') if product.created_at else '-',
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = Font(name='Arial', size=10)
            cell.fill = row_fill
            cell.alignment = center_align
            if col == 4:
                cell.number_format = '#,##0.00 ₺'
    
    col_widths = [6, 35, 15, 22, 8, 22, 22]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="urunler_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_staff_excel(request):
    """Personel listesini Excel'e aktar"""
    
    if not request.user.is_manager:
        messages.error(request, "Bu işlem için yetkiniz yok.")
        return redirect('my_offers')
    
    staff = User.objects.filter(manager=request.user).order_by('-date_joined')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Personel Listesi"
    
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', start_color='2F6FED')
    header_align = Alignment(horizontal='center', vertical='center')
    center_align = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A1:H1')
    ws['A1'] = f"PERSONEL LİSTESİ - {request.user.get_full_name() or request.user.username}"
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', start_color='1A3A6B')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:H2')
    ws['A2'] = f"Dışa Aktarma Tarihi: {timezone.now().strftime('%d.%m.%Y %H:%M')} | Toplam: {staff.count()} personel"
    ws['A2'].font = Font(name='Arial', size=10, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')
    
    headers = ['Sıra', 'Ad Soyad', 'Kullanıcı Adı', 'Email', 'Durum', 'Kayıt Tarihi', 'Son Giriş', 'Toplam Teklif']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws.row_dimensions[3].height = 30
    
    for row_idx, s in enumerate(staff, 4):
        row_fill = PatternFill('solid', start_color='F8F9FA') if row_idx % 2 == 0 else PatternFill('solid', start_color='FFFFFF')
        
        data = [
            row_idx - 3,
            s.get_full_name() or s.username,
            s.username,
            s.email,
            'Aktif' if s.is_active_user else 'Pasif',
            s.date_joined.strftime('%d.%m.%Y') if s.date_joined else '-',
            s.last_login.strftime('%d.%m.%Y %H:%M') if s.last_login else 'Hiç giriş yapmadı',
            Offer.objects.filter(user=s).count(),
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = Font(name='Arial', size=10)
            cell.fill = row_fill
            cell.alignment = center_align
            if col == 5:
                cell.font = Font(name='Arial', size=10, bold=True, 
                               color='155724' if value == 'Aktif' else '721C24')
    
    col_widths = [6, 25, 18, 28, 10, 15, 20, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="personeller_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    wb.save(response)
    return response

#PDF İçin------------------------

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

@login_required
def export_offer_excel(request, offer_id):
    """Tek teklif Excel'e aktar"""
    
    offer = get_object_or_404(Offer, id=offer_id)
    
    # Yetki kontrolü
    can_export = (
        request.user.is_superuser or request.user.is_staff or
        offer.user == request.user or
        (request.user.is_manager and offer.user.manager == request.user) or
        request.user.role == 'eczane'
    )
    
    if not can_export:
        messages.error(request, "Bu teklifi dışa aktarma yetkiniz yok.")
        return redirect('my_offers')
    
    wb = Workbook()
    
    # ========================
    # SAYFA 1: TEKLİF BİLGİLERİ
    # ========================
    ws1 = wb.active
    ws1.title = "Teklif Bilgileri"
    
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', start_color='2F6FED')
    title_fill = PatternFill('solid', start_color='1A3A6B')
    green_fill = PatternFill('solid', start_color='28A745')
    red_fill = PatternFill('solid', start_color='DC3545')
    orange_fill = PatternFill('solid', start_color='FFC107')
    label_fill = PatternFill('solid', start_color='EEF1FF')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Başlık
    ws1.merge_cells('A1:F1')
    ws1['A1'] = 'TEKLİF DETAY RAPORU'
    ws1['A1'].font = Font(name='Arial', bold=True, size=16, color='FFFFFF')
    ws1['A1'].fill = title_fill
    ws1['A1'].alignment = center_align
    ws1.row_dimensions[1].height = 40
    
    # Teklif No ve Durum
    offer_no = f"#{offer.original_offer.id if offer.original_offer else offer.id}"
    if offer.revision_number > 1:
        offer_no += f" (Rev.{offer.revision_number})"
    
    ws1.merge_cells('A2:C2')
    ws1['A2'] = f'Teklif No: {offer_no}'
    ws1['A2'].font = Font(name='Arial', bold=True, size=13, color='1A3A6B')
    ws1['A2'].alignment = left_align
    
    status_map = {'sent': 'Bekliyor', 'approved': 'Onaylandı', 'rejected': 'Reddedildi', 'revised': 'Revize Edildi'}
    status_colors = {'sent': 'FFC107', 'approved': '28A745', 'rejected': 'DC3545', 'revised': '007BFF'}
    
    ws1.merge_cells('D2:F2')
    ws1['D2'] = f'Durum: {status_map.get(offer.status, offer.status)}'
    ws1['D2'].font = Font(name='Arial', bold=True, size=13, color=status_colors.get(offer.status, '333333'))
    ws1['D2'].alignment = center_align
    ws1.row_dimensions[2].height = 30
    
    ws1.row_dimensions[3].height = 10
    
    # ---- FİRMA BİLGİLERİ ----
    ws1.merge_cells('A4:F4')
    ws1['A4'] = '🏢 FİRMA BİLGİLERİ'
    ws1['A4'].font = header_font
    ws1['A4'].fill = header_fill
    ws1['A4'].alignment = center_align
    ws1.row_dimensions[4].height = 25
    
    firma_data = [
        ['Firma Adı', offer.user.company_name or '-', 'Yetkili Kişi', offer.user.company_responsible_person or '-'],
        ['Vergi No', offer.user.company_tax_number or '-', 'Vergi Dairesi', offer.user.company_tax_office or '-'],
        ['Telefon', offer.user.company_phone or '-', 'Mobil', offer.user.company_mobile or '-'],
        ['Email', offer.user.company_responsible_email or offer.user.email or '-', 'Adres', offer.user.company_address or '-'],
        ['Teklifi Hazırlayan', offer.user.get_full_name() or offer.user.username, 'Kullanıcı Adı', offer.user.username],
    ]
    
    row = 5
    for data_row in firma_data:
        ws1.cell(row=row, column=1, value=data_row[0]).fill = label_fill
        ws1.cell(row=row, column=1).font = Font(name='Arial', bold=True, size=10)
        ws1.merge_cells(f'B{row}:C{row}')
        ws1.cell(row=row, column=2, value=data_row[1]).font = Font(name='Arial', size=10)
        ws1.cell(row=row, column=4, value=data_row[2]).fill = label_fill
        ws1.cell(row=row, column=4).font = Font(name='Arial', bold=True, size=10)
        ws1.merge_cells(f'E{row}:F{row}')
        ws1.cell(row=row, column=5, value=data_row[3]).font = Font(name='Arial', size=10)
        ws1.row_dimensions[row].height = 20
        row += 1
    
    row += 1
    
    # ---- ECZANE BİLGİLERİ ----
    ws1.merge_cells(f'A{row}:F{row}')
    ws1[f'A{row}'] = '💊 ECZANE BİLGİLERİ'
    ws1[f'A{row}'].font = header_font
    ws1[f'A{row}'].fill = PatternFill('solid', start_color='28A745')
    ws1[f'A{row}'].alignment = center_align
    ws1.row_dimensions[row].height = 25
    row += 1
    
    # Eczane bilgilerini bul
    eczane_user = None
    if offer.approved_by:
        eczane_user = offer.approved_by.manager if not offer.approved_by.is_manager else offer.approved_by
    elif offer.rejected_by:
        eczane_user = offer.rejected_by.manager if not offer.rejected_by.is_manager else offer.rejected_by
    
    eczane_data = [
        ['Eczane Adı', eczane_user.pharmacy_name if eczane_user else '-', 'Eczacı', eczane_user.pharmacist_name if eczane_user else '-'],
        ['Vergi No', eczane_user.pharmacy_tax_number if eczane_user else '-', 'Ruhsat No', eczane_user.pharmacy_license_number if eczane_user else '-'],
        ['Telefon', eczane_user.pharmacy_phone if eczane_user else '-', 'Mobil', eczane_user.pharmacy_mobile if eczane_user else '-'],
        ['Email', eczane_user.pharmacy_email if eczane_user else '-', 'Adres', eczane_user.pharmacy_address if eczane_user else '-'],
        ['Onaylayan Personel', offer.approved_by.get_full_name() if offer.approved_by else '-', 'Onay Tarihi', offer.approved_at.strftime('%d.%m.%Y %H:%M') if offer.approved_at else '-'],
        ['Reddeden Personel', offer.rejected_by.get_full_name() if offer.rejected_by else '-', 'Red Tarihi', offer.rejected_at.strftime('%d.%m.%Y %H:%M') if offer.rejected_at else '-'],
    ]
    
    for data_row in eczane_data:
        ws1.cell(row=row, column=1, value=data_row[0]).fill = PatternFill('solid', start_color='EAFFF2')
        ws1.cell(row=row, column=1).font = Font(name='Arial', bold=True, size=10)
        ws1.merge_cells(f'B{row}:C{row}')
        ws1.cell(row=row, column=2, value=data_row[1]).font = Font(name='Arial', size=10)
        ws1.cell(row=row, column=4, value=data_row[2]).fill = PatternFill('solid', start_color='EAFFF2')
        ws1.cell(row=row, column=4).font = Font(name='Arial', bold=True, size=10)
        ws1.merge_cells(f'E{row}:F{row}')
        ws1.cell(row=row, column=5, value=data_row[3]).font = Font(name='Arial', size=10)
        ws1.row_dimensions[row].height = 20
        row += 1
    
    row += 1
    
    # ---- TEKLİF ÖZETI ----
    ws1.merge_cells(f'A{row}:F{row}')
    ws1[f'A{row}'] = '📋 TEKLİF ÖZETİ'
    ws1[f'A{row}'].font = header_font
    ws1[f'A{row}'].fill = PatternFill('solid', start_color='764BA2')
    ws1[f'A{row}'].alignment = center_align
    ws1.row_dimensions[row].height = 25
    row += 1
    
    ozet_data = [
        ['Gönderilme Tarihi', offer.sent_at.strftime('%d.%m.%Y %H:%M') if offer.sent_at else '-', 'Revizyon No', str(offer.revision_number)],
        ['Fatura No', offer.invoice_number or '-', 'Fatura Tarihi', offer.invoice_date.strftime('%d.%m.%Y') if offer.invoice_date else '-'],
        ['Termin Tarihi', offer.delivery_deadline.strftime('%d.%m.%Y') if offer.delivery_deadline else '-', 'Red Nedeni', offer.reject_reason or '-'],
    ]
    
    for data_row in ozet_data:
        ws1.cell(row=row, column=1, value=data_row[0]).fill = PatternFill('solid', start_color='F3EEFF')
        ws1.cell(row=row, column=1).font = Font(name='Arial', bold=True, size=10)
        ws1.merge_cells(f'B{row}:C{row}')
        ws1.cell(row=row, column=2, value=data_row[1]).font = Font(name='Arial', size=10)
        ws1.cell(row=row, column=4, value=data_row[2]).fill = PatternFill('solid', start_color='F3EEFF')
        ws1.cell(row=row, column=4).font = Font(name='Arial', bold=True, size=10)
        ws1.merge_cells(f'E{row}:F{row}')
        ws1.cell(row=row, column=5, value=data_row[3]).font = Font(name='Arial', size=10)
        ws1.row_dimensions[row].height = 20
        row += 1
    
    # Sütun genişlikleri
    col_widths = [22, 20, 10, 22, 20, 10]
    for i, width in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = width
    
    # ========================
    # SAYFA 2: ÜRÜN DETAYLARI
    # ========================
    ws2 = wb.create_sheet("Ürün Detayları")
    
    ws2.merge_cells('A1:N1')
    ws2['A1'] = f'ÜRÜN DETAYLARI - Teklif {offer_no}'
    ws2['A1'].font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    ws2['A1'].fill = title_fill
    ws2['A1'].alignment = center_align
    ws2.row_dimensions[1].height = 35
    
    headers2 = ['Sıra', 'Ürün Adı', 'Barkod', 'Adet', 'Birim Fiyat\n(KDV Hariç)', 
                'KDV %', 'İskonto\nTipi', 'İskonto\nDeğeri', 'İskonto\nTutarı',
                'KDV Hariç\nToplam', 'KDV\nTutarı', 'KDV Dahil\nToplam',
                'Teslimat Adresi', 'Not']
    
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    ws2.row_dimensions[2].height = 35
    
    for idx, item in enumerate(offer.items.all(), 1):
        row_fill = PatternFill('solid', start_color='F8F9FA') if idx % 2 == 0 else PatternFill('solid', start_color='FFFFFF')
        
        delivery_addr = '-'
        if hasattr(item, 'delivery_address') and item.delivery_address:
            delivery_addr = f"{item.delivery_address.title} - {item.delivery_address.city}"
        
        data = [
            idx,
            item.product.name,
            item.product.barcode or '-',
            item.quantity,
            float(item.unit_price),
            item.vat_rate,
            {'none': 'Yok', 'percent': '%', 'amount': '₺'}.get(item.discount_type, '-'),
            float(item.discount_value),
            float(item.discount_amount),
            float(item.line_subtotal),
            float(item.vat_amount),
            float(item.total_price),
            delivery_addr,
            item.note or '-',
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws2.cell(row=idx+2, column=col, value=value)
            cell.font = Font(name='Arial', size=10)
            cell.fill = row_fill
            cell.alignment = center_align
            if col in [5, 9, 10, 11, 12]:
                cell.number_format = '#,##0.00 ₺'
        
        ws2.row_dimensions[idx+2].height = 20
    
    # Toplam satırı
    total_row_idx = offer.items.count() + 3
    totals = [
        ('', '', '', '', '', '', '', '', 'TOPLAM:', 
         float(offer.items_net_after_item_discounts()),
         float(offer.items_vat_after_item_discounts()),
         float(offer.final_total), '', '')
    ]
    
    for col, value in enumerate(totals[0], 1):
        cell = ws2.cell(row=total_row_idx, column=col, value=value)
        cell.font = Font(name='Arial', bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='2F6FED')
        cell.alignment = center_align
        if col in [10, 11, 12]:
            cell.number_format = '#,##0.00 ₺'
    ws2.row_dimensions[total_row_idx].height = 25
    
    col_widths2 = [6, 30, 14, 8, 18, 8, 10, 12, 14, 18, 12, 18, 25, 20]
    for i, width in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = width
    
    # ========================
    # SAYFA 3: TOPLAM ÖZETİ
    # ========================
    ws3 = wb.create_sheet("Toplam Özeti")
    
    ws3.merge_cells('A1:D1')
    ws3['A1'] = f'TOPLAM ÖZETİ - Teklif {offer_no}'
    ws3['A1'].font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    ws3['A1'].fill = title_fill
    ws3['A1'].alignment = center_align
    ws3.row_dimensions[1].height = 35
    
    summary_data = [
        ('KDV Hariç Toplam (İskonto Öncesi)', float(offer.items_subtotal_net())),
        ('Ürün İskonto Toplamı', -float(offer.total_item_discounts)),
        ('KDV Hariç Toplam (İskonto Sonrası)', float(offer.items_net_after_item_discounts())),
    ]
    
    if offer.overall_discount_type != 'none':
        label = f'Genel İskonto ({offer.overall_discount_value}{"%" if offer.overall_discount_type == "percent" else "₺"})'
        summary_data.append((label, -float(offer.overall_discount_amount)))
        summary_data.append(('Genel İskonto Sonrası KDV Hariç', float(offer.net_after_overall_discount)))
    
    summary_data.append(('KDV Toplamı', float(offer.vat_after_overall_discount if offer.overall_discount_type != 'none' else offer.items_vat_after_item_discounts())))
    summary_data.append(('SON TOPLAM (KDV DAHİL)', float(offer.final_total)))
    
    for row_idx, (label, value) in enumerate(summary_data, 2):
        is_total = label.startswith('SON TOPLAM')
        is_discount = 'İskonto' in label and not 'Sonrası' in label
        
        cell_label = ws3.cell(row=row_idx, column=1, value=label)
        cell_value = ws3.cell(row=row_idx, column=2, value=value)
        
        if is_total:
            cell_label.fill = PatternFill('solid', start_color='2F6FED')
            cell_label.font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
            cell_value.fill = PatternFill('solid', start_color='2F6FED')
            cell_value.font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
            ws3.row_dimensions[row_idx].height = 30
        elif is_discount:
            cell_label.font = Font(name='Arial', size=11, color='DC3545')
            cell_value.font = Font(name='Arial', size=11, color='DC3545')
        else:
            cell_label.font = Font(name='Arial', size=11)
            cell_value.font = Font(name='Arial', size=11)
        
        cell_label.alignment = left_align
        cell_value.alignment = center_align
        cell_value.number_format = '#,##0.00 ₺'
        ws3.row_dimensions[row_idx].height = 22
    
    ws3.column_dimensions['A'].width = 40
    ws3.column_dimensions['B'].width = 20
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="teklif_{offer.id}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_offer_pdf(request, offer_id):
    """Tek teklif PDF'e aktar"""
    
    offer = get_object_or_404(Offer, id=offer_id)
    
    can_export = (
        request.user.is_superuser or request.user.is_staff or
        offer.user == request.user or
        (request.user.is_manager and offer.user.manager == request.user) or
        request.user.role == 'eczane'
    )
    
    if not can_export:
        messages.error(request, "Bu teklifi dışa aktarma yetkiniz yok.")
        return redirect('my_offers')
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="teklif_{offer.id}_{timezone.now().strftime("%Y%m%d")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4, 
                           rightMargin=1.5*cm, leftMargin=1.5*cm,
                           topMargin=1.5*cm, bottomMargin=1.5*cm)
    
    styles = getSampleStyleSheet()
    story = []
    
    # Özel stiller
    title_style = ParagraphStyle('title', parent=styles['Normal'],
                                fontSize=18, fontName='Helvetica-Bold',
                                textColor=colors.HexColor('#1A3A6B'),
                                alignment=TA_CENTER, spaceAfter=8)
    
    h1_style = ParagraphStyle('h1', parent=styles['Normal'],
                             fontSize=12, fontName='Helvetica-Bold',
                             textColor=colors.white,
                             backColor=colors.HexColor('#2F6FED'),
                             alignment=TA_CENTER, spaceBefore=10, spaceAfter=4,
                             leftPadding=6, rightPadding=6, topPadding=4, bottomPadding=4)
    
    h2_style = ParagraphStyle('h2', parent=styles['Normal'],
                             fontSize=11, fontName='Helvetica-Bold',
                             textColor=colors.white,
                             backColor=colors.HexColor('#28A745'),
                             alignment=TA_CENTER, spaceBefore=8, spaceAfter=4,
                             leftPadding=6, rightPadding=6, topPadding=4, bottomPadding=4)
    
    label_style = ParagraphStyle('label', parent=styles['Normal'],
                                fontSize=9, fontName='Helvetica-Bold',
                                textColor=colors.HexColor('#333333'))
    
    value_style = ParagraphStyle('value', parent=styles['Normal'],
                                fontSize=9, fontName='Helvetica',
                                textColor=colors.HexColor('#555555'))
    
    normal_style = ParagraphStyle('normal', parent=styles['Normal'],
                                 fontSize=9, fontName='Helvetica')
    
    # Başlık
    offer_no = f"#{offer.original_offer.id if offer.original_offer else offer.id}"
    if offer.revision_number > 1:
        offer_no += f" (Rev.{offer.revision_number})"
    
    story.append(Paragraph(f"TEKLIF DETAY RAPORU", title_style))
    
    status_map = {'sent': 'Bekliyor', 'approved': 'Onaylandi', 'rejected': 'Reddedildi', 'revised': 'Revize Edildi'}
    status_colors_pdf = {'sent': '#856404', 'approved': '#155724', 'rejected': '#721C24', 'revised': '#004085'}
    
    status_color = status_colors_pdf.get(offer.status, '#333333')
    story.append(Paragraph(
        f'<font color="{status_color}"><b>Teklif No: {offer_no} | Durum: {status_map.get(offer.status, offer.status)}</b></font>',
        ParagraphStyle('status', parent=styles['Normal'], fontSize=11, alignment=TA_CENTER, spaceAfter=10)
    ))
    
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#2F6FED')))
    story.append(Spacer(1, 0.3*cm))
    
    # Firma Bilgileri
    story.append(Paragraph("FIRMA BILGILERI", h1_style))
    
    firma_table_data = [
        [Paragraph('<b>Firma Adi</b>', label_style), Paragraph(offer.user.company_name or '-', value_style),
         Paragraph('<b>Yetkili</b>', label_style), Paragraph(offer.user.company_responsible_person or '-', value_style)],
        [Paragraph('<b>Vergi No</b>', label_style), Paragraph(offer.user.company_tax_number or '-', value_style),
         Paragraph('<b>Vergi Dairesi</b>', label_style), Paragraph(offer.user.company_tax_office or '-', value_style)],
        [Paragraph('<b>Telefon</b>', label_style), Paragraph(offer.user.company_phone or '-', value_style),
         Paragraph('<b>Mobil</b>', label_style), Paragraph(offer.user.company_mobile or '-', value_style)],
        [Paragraph('<b>Email</b>', label_style), Paragraph(offer.user.email or '-', value_style),
         Paragraph('<b>Adres</b>', label_style), Paragraph(offer.user.company_address or '-', value_style)],
        [Paragraph('<b>Hazirlayan</b>', label_style), Paragraph(offer.user.get_full_name() or offer.user.username, value_style),
         Paragraph('<b>Kullanici Adi</b>', label_style), Paragraph(offer.user.username, value_style)],
    ]
    
    firma_table = Table(firma_table_data, colWidths=[3.5*cm, 6*cm, 3.5*cm, 5.5*cm])
    firma_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#EEF1FF')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#EEF1FF')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(firma_table)
    story.append(Spacer(1, 0.3*cm))
    
    # Eczane Bilgileri
    story.append(Paragraph("ECZANE BILGILERI", h2_style))
    
    eczane_user = None
    if offer.approved_by:
        eczane_user = offer.approved_by.manager if not offer.approved_by.is_manager else offer.approved_by
    elif offer.rejected_by:
        eczane_user = offer.rejected_by.manager if not offer.rejected_by.is_manager else offer.rejected_by
    
    eczane_table_data = [
        [Paragraph('<b>Eczane Adi</b>', label_style), Paragraph(eczane_user.pharmacy_name if eczane_user else '-', value_style),
         Paragraph('<b>Eczaci</b>', label_style), Paragraph(eczane_user.pharmacist_name if eczane_user else '-', value_style)],
        [Paragraph('<b>Vergi No</b>', label_style), Paragraph(eczane_user.pharmacy_tax_number if eczane_user else '-', value_style),
         Paragraph('<b>Ruhsat No</b>', label_style), Paragraph(eczane_user.pharmacy_license_number if eczane_user else '-', value_style)],
        [Paragraph('<b>Telefon</b>', label_style), Paragraph(eczane_user.pharmacy_phone if eczane_user else '-', value_style),
         Paragraph('<b>Email</b>', label_style), Paragraph(eczane_user.pharmacy_email if eczane_user else '-', value_style)],
        [Paragraph('<b>Adres</b>', label_style), Paragraph(eczane_user.pharmacy_address if eczane_user else '-', value_style),
         Paragraph('<b>Onaylayan</b>', label_style), Paragraph(offer.approved_by.get_full_name() if offer.approved_by else '-', value_style)],
        [Paragraph('<b>Onay Tarihi</b>', label_style), Paragraph(offer.approved_at.strftime('%d.%m.%Y %H:%M') if offer.approved_at else '-', value_style),
         Paragraph('<b>Termin Tarihi</b>', label_style), Paragraph(offer.delivery_deadline.strftime('%d.%m.%Y') if offer.delivery_deadline else '-', value_style)],
    ]
    
    eczane_table = Table(eczane_table_data, colWidths=[3.5*cm, 6*cm, 3.5*cm, 5.5*cm])
    eczane_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#EAFFF2')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#EAFFF2')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(eczane_table)
    story.append(Spacer(1, 0.3*cm))
    
    # Fatura Bilgileri
    if offer.invoice_number or offer.invoice_date or offer.delivery_deadline:
        story.append(Paragraph("FATURA BILGILERI", ParagraphStyle('h3', parent=styles['Normal'],
                              fontSize=11, fontName='Helvetica-Bold', textColor=colors.white,
                              backColor=colors.HexColor('#764BA2'), alignment=TA_CENTER,
                              spaceBefore=8, spaceAfter=4, leftPadding=6, topPadding=4, bottomPadding=4)))
        
        fatura_data = [
            [Paragraph('<b>Fatura No</b>', label_style), Paragraph(offer.invoice_number or '-', value_style),
             Paragraph('<b>Fatura Tarihi</b>', label_style), Paragraph(offer.invoice_date.strftime('%d.%m.%Y') if offer.invoice_date else '-', value_style)],
            [Paragraph('<b>Termin Tarihi</b>', label_style), Paragraph(offer.delivery_deadline.strftime('%d.%m.%Y') if offer.delivery_deadline else '-', value_style),
             Paragraph('<b>Gonderi Tarihi</b>', label_style), Paragraph(offer.sent_at.strftime('%d.%m.%Y %H:%M') if offer.sent_at else '-', value_style)],
        ]
        
        fatura_table = Table(fatura_data, colWidths=[3.5*cm, 6*cm, 3.5*cm, 5.5*cm])
        fatura_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F3EEFF')),
            ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#F3EEFF')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(fatura_table)
        story.append(Spacer(1, 0.3*cm))
    
    # Ürün Detayları
    story.append(Paragraph("URUN DETAYLARI", h1_style))
    
    urun_headers = [
        Paragraph('<b>Sira</b>', label_style),
        Paragraph('<b>Urun Adi</b>', label_style),
        Paragraph('<b>Adet</b>', label_style),
        Paragraph('<b>Birim\nFiyat</b>', label_style),
        Paragraph('<b>KDV\n%</b>', label_style),
        Paragraph('<b>Iskonto</b>', label_style),
        Paragraph('<b>KDV\nHaric</b>', label_style),
        Paragraph('<b>KDV</b>', label_style),
        Paragraph('<b>Toplam</b>', label_style),
        Paragraph('<b>Teslimat\nAdresi</b>', label_style),
    ]
    
    urun_data = [urun_headers]
    
    for idx, item in enumerate(offer.items.all(), 1):
        delivery_addr = '-'
        if hasattr(item, 'delivery_address') and item.delivery_address:
            delivery_addr = f"{item.delivery_address.title}"
        
        discount_str = '-'
        if item.discount_type != 'none':
            discount_str = f"{item.discount_value}{'%' if item.discount_type == 'percent' else 'TL'} (-{float(item.discount_amount):.2f})"
        
        row_data = [
            Paragraph(str(idx), normal_style),
            Paragraph(item.product.name, normal_style),
            Paragraph(str(item.quantity), normal_style),
            Paragraph(f"{float(item.unit_price):.2f}", normal_style),
            Paragraph(f"%{item.vat_rate}", normal_style),
            Paragraph(discount_str, normal_style),
            Paragraph(f"{float(item.line_subtotal):.2f}", normal_style),
            Paragraph(f"{float(item.vat_amount):.2f}", normal_style),
            Paragraph(f"{float(item.total_price):.2f}", normal_style),
            Paragraph(delivery_addr, normal_style),
        ]
        urun_data.append(row_data)
    
    # Toplam satırı
    urun_data.append([
        Paragraph('', normal_style),
        Paragraph('<b>TOPLAM</b>', label_style),
        Paragraph('', normal_style),
        Paragraph('', normal_style),
        Paragraph('', normal_style),
        Paragraph('', normal_style),
        Paragraph(f'<b>{float(offer.items_net_after_item_discounts()):.2f}</b>', label_style),
        Paragraph(f'<b>{float(offer.items_vat_after_item_discounts()):.2f}</b>', label_style),
        Paragraph(f'<b>{float(offer.final_total):.2f}</b>', label_style),
        Paragraph('', normal_style),
    ])
    
    urun_table = Table(urun_data, colWidths=[1*cm, 5*cm, 1.2*cm, 1.8*cm, 1.2*cm, 2.5*cm, 2*cm, 1.8*cm, 2*cm, 2.5*cm])
    urun_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F6FED')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#EEF1FF')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8F9FA')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(urun_table)
    story.append(Spacer(1, 0.5*cm))
    
    # Toplam Özeti
    story.append(Paragraph("TOPLAM OZETI", h1_style))
    
    toplam_data = [
        ['KDV Haric Toplam (Iskonto Oncesi)', f"{float(offer.items_subtotal_net()):.2f} TL"],
        ['Urun Iskonto Toplami', f"-{float(offer.total_item_discounts):.2f} TL"],
        ['KDV Haric Toplam (Iskonto Sonrasi)', f"{float(offer.items_net_after_item_discounts()):.2f} TL"],
    ]
    
    if offer.overall_discount_type != 'none':
        toplam_data.append([f'Genel Iskonto', f"-{float(offer.overall_discount_amount):.2f} TL"])
        toplam_data.append(['Genel Iskonto Sonrasi KDV Haric', f"{float(offer.net_after_overall_discount):.2f} TL"])
    
    kdv = offer.vat_after_overall_discount if offer.overall_discount_type != 'none' else offer.items_vat_after_item_discounts()
    toplam_data.append(['KDV Toplami', f"{float(kdv):.2f} TL"])
    toplam_data.append(['SON TOPLAM (KDV DAHIL)', f"{float(offer.final_total):.2f} TL"])
    
    toplam_table_data = []
    for label, value in toplam_data:
        is_total = label.startswith('SON TOPLAM')
        is_discount = 'Iskonto' in label and 'Sonrasi' not in label
        toplam_table_data.append([Paragraph(f'<b>{label}</b>' if is_total else label, 
                                           ParagraphStyle('tl', parent=styles['Normal'], fontSize=10,
                                                         textColor=colors.HexColor('#DC3545') if is_discount else colors.HexColor('#333333'))),
                                  Paragraph(f'<b>{value}</b>' if is_total else value,
                                           ParagraphStyle('tv', parent=styles['Normal'], fontSize=10, alignment=TA_RIGHT,
                                                         textColor=colors.HexColor('#DC3545') if is_discount else colors.HexColor('#333333')))])
    
    toplam_table = Table(toplam_table_data, colWidths=[13*cm, 5.5*cm])
    toplam_table.setStyle(TableStyle([
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#2F6FED')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(toplam_table)
    
    # Red nedeni varsa
    if offer.reject_reason:
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph("RED NEDENI", ParagraphStyle('red', parent=styles['Normal'],
                              fontSize=11, fontName='Helvetica-Bold', textColor=colors.white,
                              backColor=colors.HexColor('#DC3545'), alignment=TA_CENTER,
                              spaceBefore=8, spaceAfter=4, leftPadding=6, topPadding=4, bottomPadding=4)))
        story.append(Paragraph(offer.reject_reason, ParagraphStyle('rednote', parent=styles['Normal'],
                              fontSize=10, textColor=colors.HexColor('#721C24'),
                              backColor=colors.HexColor('#FFF2F2'), leftPadding=8, topPadding=6, bottomPadding=6)))
    
    doc.build(story)
    return response